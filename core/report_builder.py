from datetime import datetime
from collections import defaultdict
from core.enums import Gender

class ReportBuilder:
    @staticmethod
    def _calculate_age_in_months(dob_str, date_given_str):
        try:
            dob = datetime.strptime(dob_str, "%Y-%m-%d").date()
            date_given = datetime.strptime(date_given_str, "%Y-%m-%d").date()
            days = (date_given - dob).days
            months = days / 30.44
            return months
        except Exception:
            return 0

    @staticmethod
    def _map_vax_name(record):
        vax = record["vax_name"]
        if vax == "HB0":
            if record["date_given"] == record["dob"]:
                return "HB0_24h"
            return "HB0_BCG"
        if vax == "Pneumo1": return "PCV1"
        if vax == "Pneumo2": return "PCV2"
        if vax == "Pneumo3_NewOnly": return "PCV3"
        if vax == "Pneumo_Final":
            return "PCV4" if record.get("has_pneumo3", 0) == 1 else "PCV3"
        return vax

    @staticmethod
    def _process_data(stats_data, expected_vaxes):
        """
        Groups data by vaccine -> { 
            "0-11m": {"M": 0, "F": 0, "Total": 0},
            "12-59m": {"M": 0, "F": 0, "Total": 0},
            "60m+": {"M": 0, "F": 0, "Total": 0},
        }
        """
        results = {}
        for row in expected_vaxes:
            v_name = row["db_name"]
            if v_name not in results:
                results[v_name] = {
                    "0-11m": {"M": 0, "F": 0, "Total": 0},
                    "12-59m": {"M": 0, "F": 0, "Total": 0},
                    "60m+": {"M": 0, "F": 0, "Total": 0},
                }

        for record in stats_data:
            mapped_vax = ReportBuilder._map_vax_name(record)
            
            # Match to our structure
            vax_key = None
            for row in expected_vaxes:
                if row.get("db_name") == mapped_vax:
                    vax_key = row["db_name"]
                    break
            
            if not vax_key or vax_key not in results:
                continue

            sexe = record["sexe"]
            age_months = ReportBuilder._calculate_age_in_months(record["dob"], record["date_given"])
            
            if age_months < 12:
                age_group = "0-11m"
            elif age_months < 60:
                age_group = "12-59m"
            else:
                age_group = "60m+"
                
            sexe_key = "M" if sexe == Gender.MALE.value else "F"

            results[vax_key][age_group][sexe_key] += 1
            results[vax_key][age_group]["Total"] += 1
            
        return results

    @staticmethod
    def generate_fiche_html(stats_data, year, month, center_name, center_type, is_daily=False, specific_date_str=None):
        """
        Generates the HTML mimicking 'Fiche mensuelle (1).xlsx'.
        stats_data: list of dicts from get_detailed_export_stats
        """
        
        # Define the exact structure we want to render based on the CSV extract of the Excel file
        table_structure = [
            {"label": "Hep.B", "is_header": True},
            {"label": "Fait dans les 24 heures", "db_name": "HB0_24h", "has_0_11": True, "has_12_59": False},
            {"label": "Fait avec le BCG", "db_name": "HB0_BCG", "has_0_11": True, "has_12_59": False},
            
            {"label": "BCG", "is_header": True},
            {"label": "Doses administrées", "db_name": "BCG", "has_0_11": True, "has_12_59": True},
            
            {"label": "VPO.0", "is_header": True},
            {"label": "Doses administrées", "db_name": "VPO0", "has_0_11": True, "has_12_59": False},
            
            {"label": "VPO", "is_header": True},
            {"label": "1ère prise", "db_name": "VPO1", "has_0_11": True, "has_12_59": True},
            {"label": "2ème prise", "db_name": "VPO2", "has_0_11": True, "has_12_59": True},
            {"label": "3ème prise", "db_name": "VPO3", "has_0_11": True, "has_12_59": True},
            
            {"label": "VPI", "is_header": True},
            {"label": "1ère prise", "db_name": "VPI", "has_0_11": True, "has_12_59": True},
            {"label": "2ème prise", "db_name": "VPI2", "has_0_11": True, "has_12_59": True},
            
            {"label": "DTC-Hib-Hep.B (Penta)", "is_header": True},
            {"label": "1ère prise", "db_name": "Pentavalent 1", "has_0_11": True, "has_12_59": True},
            {"label": "2ème prise", "db_name": "Pentavalent 2", "has_0_11": True, "has_12_59": True},
            {"label": "3ème prise", "db_name": "Pentavalent 3", "has_0_11": True, "has_12_59": True},
            
            {"label": "Rotavirus", "is_header": True},
            {"label": "1ère prise", "db_name": "Rota1", "has_0_11": True, "has_12_59": False},
            {"label": "2ème prise", "db_name": "Rota2", "has_0_11": True, "has_12_59": False},
            {"label": "3ème prise", "db_name": "Rota3", "has_0_11": True, "has_12_59": False},
            
            {"label": "Pneumocoque", "is_header": True},
            {"label": "1ère prise", "db_name": "PCV1", "has_0_11": True, "has_12_59": True},
            {"label": "2ème prise", "db_name": "PCV2", "has_0_11": True, "has_12_59": True},
            {"label": "3ème prise", "db_name": "PCV3", "has_0_11": True, "has_12_59": True},
            {"label": "4ème prise", "db_name": "PCV4", "has_0_11": False, "has_12_59": True},
            
            {"label": "RR", "is_header": True},
            {"label": "1ère dose (9 mois)", "db_name": "RR1", "has_0_11": True, "has_12_59": True},
            {"label": "2éme dose (18 mois)", "db_name": "RR2", "has_0_11": False, "has_12_59": True},
            
            {"label": "Rappels", "is_header": True},
            {"label": "1er rappel VPO (18 mois)", "db_name": "VPO (18 Mois)", "has_0_11": False, "has_12_59": True},
            {"label": "2ème rappel VPO (5 ans)", "db_name": "VPO (5 Ans)", "has_0_11": False, "has_5y": True},
            {"label": "1er rappel DTC (18 mois)", "db_name": "Rappel DTC1", "has_0_11": False, "has_12_59": True},
            {"label": "2ème rappel DTC (5 ans)", "db_name": "Rappel DTC2", "has_0_11": False, "has_5y": True},
        ]

        expected_vaxes = [t for t in table_structure if not t.get("is_header")]
        
        # Inject custom/archived vaccines dynamically
        known_db_names = {t["db_name"] for t in expected_vaxes}
        custom_vaxes = set()
        for rec in stats_data:
            mapped_name = ReportBuilder._map_vax_name(rec)
            if mapped_name not in known_db_names:
                custom_vaxes.add(mapped_name)
                
        if custom_vaxes:
            table_structure.append({"label": "Autres Vaccins (Archives)", "is_header": True})
            for cv in sorted(list(custom_vaxes)):
                new_entry = {"label": cv, "db_name": cv, "has_0_11": True, "has_12_59": True, "has_5y": True}
                table_structure.append(new_entry)
                expected_vaxes.append(new_entry)

        processed = ReportBuilder._process_data(stats_data, expected_vaxes)
        
        # Subtitle logic
        if is_daily and specific_date_str:
            d_fmt = datetime.strptime(specific_date_str, "%Y-%m-%d").strftime("%d/%m/%Y")
            date_subtitle = f"Journée du : {d_fmt}"
        else:
            date_subtitle = f"Mois : {month} | Année : {year}"

        # HTML generation
        html = f"""
        <html>
        <head>
        <style>
            @page {{ margin: 10mm; size: landscape; }}
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; font-size: 11pt; color: #333; margin: 0; padding: 0; }}
            h1 {{ text-align: center; font-size: 20pt; margin-bottom: 5px; color: #2c3e50; text-transform: uppercase; letter-spacing: 1px; }}
            .header-info {{ 
                text-align: center; font-size: 11pt; margin-bottom: 20px; font-weight: normal; color: #555;
                background-color: #ecf0f1; border-radius: 8px; padding: 10px; max-width: 600px; margin-left: auto; margin-right: auto;
                border-left: 5px solid #3498db;
            }}
            .header-info strong {{ color: #2c3e50; }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 15px; box-shadow: 0 2px 8px rgba(0,0,0,0.1); border-radius: 8px; overflow: hidden; }}
            th, td {{ border: 1px solid #ddd; padding: 8px 12px; text-align: center; vertical-align: middle; }}
            th {{ background-color: #34495e; color: #fff; font-weight: bold; font-size: 10pt; text-transform: uppercase; }}
            .text-left {{ text-align: left; padding-left: 15px; font-weight: 500; color: #34495e; }}
            .header-row td {{ background-color: #ecf0f1; color: #2c3e50; font-weight: bold; text-align: left; padding-left: 10px; font-size: 11pt; border-left: 4px solid #3498db; }}
            .totals td {{ font-weight: bold; background-color: #fdfefe; }}
            td.count-cell {{ font-size: 11pt; color: #555; }}
            td.total-cell {{ font-weight: bold; color: #e74c3c; font-size: 12pt; background-color: #fdfefe; }}
            tr:nth-child(even) {{ background-color: #f9f9f9; }}
        </style>
        </head>
        <body>
            <h1>FICHE {'QUOTIDIENNE' if is_daily else 'MENSUELLE'} DE VACCINATION</h1>
            <div class="header-info">
                <strong>{date_subtitle}</strong><br>
                Etablissement : <strong>{center_name}</strong> ({center_type})
            </div>
            
            <table>
                <tr>
                    <th rowspan="2" style="width: 30%;">Vaccination</th>
                    <th colspan="3">0 - 11 mois</th>
                    <th colspan="3">12 - 59 mois</th>
                    <th colspan="3">≥ 5 ans</th>
                    <th rowspan="2" style="width: 10%; background-color: #2980b9;">Total Général</th>
                </tr>
                <tr>
                    <th>M</th><th>F</th><th>T</th>
                    <th>M</th><th>F</th><th>T</th>
                    <th>M</th><th>F</th><th>T</th>
                </tr>
        """
        
        for row in table_structure:
            if row.get("is_header"):
                html += f"<tr class='header-row'><td colspan='11'>{row['label']}</td></tr>"
            else:
                db_name = row["db_name"]
                data = processed.get(db_name, {
                    "0-11m": {"M": 0, "F": 0, "Total": 0},
                    "12-59m": {"M": 0, "F": 0, "Total": 0},
                    "60m+": {"M": 0, "F": 0, "Total": 0}
                })
                
                has_12_59 = row.get("has_12_59", False)
                has_5y = row.get("has_5y", False)
                
                # Render 0-11m
                c_0_m = data["0-11m"]["M"] if row.get("has_0_11") else ""
                c_0_f = data["0-11m"]["F"] if row.get("has_0_11") else ""
                c_0_t = data["0-11m"]["Total"] if row.get("has_0_11") else ""
                
                # Render 12-59m
                c_12_m = data["12-59m"]["M"] if has_12_59 else ""
                c_12_f = data["12-59m"]["F"] if has_12_59 else ""
                c_12_t = data["12-59m"]["Total"] if has_12_59 else ""
                
                # Render 5y+
                c_5_m = data["60m+"]["M"] if has_5y else ""
                c_5_f = data["60m+"]["F"] if has_5y else ""
                c_5_t = data["60m+"]["Total"] if has_5y else ""
                
                grand_total = data["0-11m"]["Total"] + data["12-59m"]["Total"] + data["60m+"]["Total"]
                grand_total_display = grand_total if grand_total > 0 else ""
                
                html += f"""
                <tr>
                    <td class="text-left">{row['label']}</td>
                    <td class="count-cell">{c_0_m}</td><td class="count-cell">{c_0_f}</td><td style="font-weight:bold; color:#2980b9;">{c_0_t}</td>
                    <td class="count-cell">{c_12_m}</td><td class="count-cell">{c_12_f}</td><td style="font-weight:bold; color:#2980b9;">{c_12_t}</td>
                    <td class="count-cell">{c_5_m}</td><td class="count-cell">{c_5_f}</td><td style="font-weight:bold; color:#2980b9;">{c_5_t}</td>
                    <td class="total-cell">{grand_total_display}</td>
                </tr>
                """
                
        html += """
            </table>
        </body>
        </html>
        """
        return html

    @staticmethod
    def generate_daily_breakdown_html(engine, year, month, center_name, center_type):
        """
        Generates a crosstab matrix PDF for the entire month.
        Rows: Vaccines
        Columns: Days 1..31
        """
        import calendar
        
        y_int = int(year)
        m_int = int(month)
        _, num_days = calendar.monthrange(y_int, m_int)
        
        # Row configuration
        vax_rows = [
            ("HB0_24h", "HB0 (Dans les 24h)"),
            ("HB0_BCG", "HB0 (Plus tard)"),
            ("BCG", "BCG"),
            ("VPO0", "VPO 0"),
            ("VPO1", "VPO 1"),
            ("VPO2", "VPO 2"),
            ("VPO3", "VPO 3"),
            ("VPI", "VPI 1"),
            ("VPI2", "VPI 2"),
            ("Pentavalent 1", "Penta 1"),
            ("Pentavalent 2", "Penta 2"),
            ("Pentavalent 3", "Penta 3"),
            ("Rota1", "Rota 1"),
            ("Rota2", "Rota 2"),
            ("Rota3", "Rota 3"),
            ("PCV1", "Pneumo 1"),
            ("PCV2", "Pneumo 2"),
            ("PCV3", "Pneumo 3"),
            ("PCV4", "Pneumo 4"),
            ("RR1", "RR 1 (9 mois)"),
            ("RR2", "RR 2 (18 mois)"),
            ("VPO (18 Mois)", "Rappel VPO (18m)"),
            ("VPO (5 Ans)", "Rappel VPO (5 ans)"),
            ("Rappel DTC1", "Rappel DTC 1"),
            ("Rappel DTC2", "Rappel DTC 2"),
        ]
        
        year_month_str = f"{year}-{month}%"
        stats_data = engine.db.get_detailed_export_stats(year_month_str)
        
        # Initialize grid
        grid = {db_name: {d: 0 for d in range(1, num_days + 1)} for db_name, _ in vax_rows}
        day_totals = {d: 0 for d in range(1, num_days + 1)}
        
        for record in stats_data:
            mapped_name = ReportBuilder._map_vax_name(record)
            day = int(record["date_given"][-2:]) # "2024-03-15" -> 15
            
            # If mapped_name isn't in grid, it's an archived/custom vaccine. Add it dynamically.
            if mapped_name not in grid:
                vax_rows.append((mapped_name, mapped_name))
                grid[mapped_name] = {d: 0 for d in range(1, num_days + 1)}
            
            if 1 <= day <= num_days:
                grid[mapped_name][day] += 1
                day_totals[day] += 1
                
        
        # Filter days
        active_days = [d for d in range(1, num_days + 1) if day_totals[d] > 0]
        
        # Build HTML
        html = f"""
        <html>
        <head>
        <style>
            @page {{ margin: 10mm; size: landscape; }}
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; font-size: 8pt; color: #333; margin: 0; padding: 0; }}
            h1 {{ text-align: center; font-size: 16pt; margin-bottom: 5px; color: #2c3e50; text-transform: uppercase; letter-spacing: 1px; }}
            .header-info {{ 
                text-align: center; font-size: 10pt; margin-bottom: 15px; color: #555;
                background-color: #ecf0f1; border-radius: 6px; padding: 8px; max-width: 600px; margin-left: auto; margin-right: auto;
                border-left: 5px solid #e74c3c;
            }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; box-shadow: 0 1px 4px rgba(0,0,0,0.1); border-radius: 4px; overflow: hidden; }}
            th, td {{ border: 1px solid #ccc; padding: 4px; text-align: center; vertical-align: middle; }}
            th {{ background-color: #34495e; color: #fff; font-weight: bold; font-size: 8pt; }}
            .text-left {{ text-align: left; padding-left: 10px; font-weight: bold; color: #2c3e50; font-size: 8pt; white-space: nowrap; }}
            .zero-cell {{ color: #ccc; }}
            .count-cell {{ font-weight: bold; color: #2980b9; }}
            .row-total {{ font-weight: bold; background-color: #fdfefe; color: #e74c3c; }}
            .footer-row td {{ background-color: #ecf0f1; font-weight: bold; color: #2c3e50; }}
            tr:nth-child(even) {{ background-color: #fafafa; }}
            tr:hover {{ background-color: #e8f4f8; }}
        </style>
        </head>
        <body>
            <h1>REGISTRE JOURNALIER DE VACCINATION</h1>
            <div class="header-info">
                Mois : <strong>{month}/{year}</strong><br>
                Etablissement : <strong>{center_name}</strong> ({center_type})
            </div>
            
            <table>
                <tr>
                    <th style="width: 15%;">Vaccin</th>
        """
        for d in active_days:
            html += f"<th>{d:02d}/{m_int:02d}</th>"
            
        html += """
                    <th style="width: 5%; background-color: #2980b9;">Total</th>
                </tr>
        """
        
        for db_name, label in vax_rows:
            html += f"<tr><td class='text-left'>{label}</td>"
            row_tot = sum(grid[db_name][d] for d in range(1, num_days + 1))
            for d in active_days:
                val = grid[db_name][d]
                cls = "count-cell" if val > 0 else "zero-cell"
                display_val = val if val > 0 else "-"
                html += f"<td class='{cls}'>{display_val}</td>"
            html += f"<td class='row-total'>{row_tot if row_tot > 0 else '-'}</td></tr>"
            
        # Footer totals
        html += "<tr class='footer-row'><td class='text-left'>Total Doses</td>"
        grand_total = sum(day_totals[d] for d in range(1, num_days + 1))
        for d in active_days:
            val = day_totals[d]
            html += f"<td>{val if val > 0 else '-'}</td>"
        html += f"<td style='color: #e74c3c;'>{grand_total if grand_total > 0 else '-'}</td></tr>"
        
        html += """
            </table>
        </body>
        </html>
        """
        return html

    @staticmethod
    def generate_fiche_excel(stats_data, year, month, center_name, center_type):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Mensuelle_{month}_{year}"
        
        # Define styles
        title_font = Font(name="Segoe UI", size=16, bold=True, color="2C3E50")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        sub_header_font = Font(name="Segoe UI", size=10, bold=True, color="2C3E50")
        normal_font = Font(name="Segoe UI", size=10)
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        total_font = Font(name="Segoe UI", size=11, bold=True, color="E74C3C")
        
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        sub_header_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        striped_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        
        thin_border = Border(left=Side(style='thin', color="DDDDDD"), 
                             right=Side(style='thin', color="DDDDDD"), 
                             top=Side(style='thin', color="DDDDDD"), 
                             bottom=Side(style='thin', color="DDDDDD"))
                             
        # Titles
        ws.merge_cells("A1:K1")
        ws["A1"] = "FICHE MENSUELLE DE VACCINATION"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align
        
        ws.merge_cells("A2:K2")
        ws["A2"] = f"Mois : {month} | Année : {year}    -    Etablissement : {center_name} ({center_type})"
        ws["A2"].font = sub_header_font
        ws["A2"].alignment = center_align
        
        # Headers Row 1
        headers_r1 = [
            ("A4:A5", "Vaccination"),
            ("B4:D4", "0 - 11 mois"),
            ("E4:G4", "12 - 59 mois"),
            ("H4:J4", "≥ 5 ans"),
            ("K4:K5", "Total Général")
        ]
        
        for rng, label in headers_r1:
            ws.merge_cells(rng)
            cell = ws[rng.split(":")[0]]
            cell.value = label
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = thin_border
            
        # Headers Row 2 (M, F, T)
        for col in ["B", "E", "H"]: ws[f"{col}5"] = "M"
        for col in ["C", "F", "I"]: ws[f"{col}5"] = "F"
        for col in ["D", "G", "J"]: ws[f"{col}5"] = "T"
        
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
            ws[f"{col}5"].font = header_font
            ws[f"{col}5"].alignment = center_align
            ws[f"{col}5"].fill = header_fill
            ws[f"{col}5"].border = thin_border
            
        # Process Data same as HTML
        table_structure = [
            {"label": "Hep.B", "is_header": True},
            {"label": "Fait dans les 24 heures", "db_name": "HB0_24h", "has_0_11": True, "has_12_59": False},
            {"label": "Fait avec le BCG", "db_name": "HB0_BCG", "has_0_11": True, "has_12_59": False},
            {"label": "BCG", "is_header": True},
            {"label": "Doses administrées", "db_name": "BCG", "has_0_11": True, "has_12_59": True},
            {"label": "VPO.0", "is_header": True},
            {"label": "Doses administrées", "db_name": "VPO0", "has_0_11": True, "has_12_59": False},
            {"label": "VPO", "is_header": True},
            {"label": "1ère prise", "db_name": "VPO1", "has_0_11": True, "has_12_59": True},
            {"label": "2ème prise", "db_name": "VPO2", "has_0_11": True, "has_12_59": True},
            {"label": "3ème prise", "db_name": "VPO3", "has_0_11": True, "has_12_59": True},
            {"label": "VPI", "is_header": True},
            {"label": "1ère prise", "db_name": "VPI", "has_0_11": True, "has_12_59": True},
            {"label": "2ème prise", "db_name": "VPI2", "has_0_11": True, "has_12_59": True},
            {"label": "DTC-Hib-Hep.B (Penta)", "is_header": True},
            {"label": "1ère prise", "db_name": "Pentavalent 1", "has_0_11": True, "has_12_59": True},
            {"label": "2ème prise", "db_name": "Pentavalent 2", "has_0_11": True, "has_12_59": True},
            {"label": "3ème prise", "db_name": "Pentavalent 3", "has_0_11": True, "has_12_59": True},
            {"label": "Rotavirus", "is_header": True},
            {"label": "1ère prise", "db_name": "Rota1", "has_0_11": True, "has_12_59": False},
            {"label": "2ème prise", "db_name": "Rota2", "has_0_11": True, "has_12_59": False},
            {"label": "3ème prise", "db_name": "Rota3", "has_0_11": True, "has_12_59": False},
            {"label": "Pneumocoque", "is_header": True},
            {"label": "1ère prise", "db_name": "PCV1", "has_0_11": True, "has_12_59": True},
            {"label": "2ème prise", "db_name": "PCV2", "has_0_11": True, "has_12_59": True},
            {"label": "3ème prise", "db_name": "PCV3", "has_0_11": True, "has_12_59": True},
            {"label": "4ème prise", "db_name": "PCV4", "has_0_11": False, "has_12_59": True},
            {"label": "RR", "is_header": True},
            {"label": "1ère dose (9 mois)", "db_name": "RR1", "has_0_11": True, "has_12_59": True},
            {"label": "2éme dose (18 mois)", "db_name": "RR2", "has_0_11": False, "has_12_59": True},
            {"label": "Rappels", "is_header": True},
            {"label": "1er rappel VPO (18 mois)", "db_name": "VPO (18 Mois)", "has_0_11": False, "has_12_59": True},
            {"label": "2ème rappel VPO (5 ans)", "db_name": "VPO (5 Ans)", "has_0_11": False, "has_5y": True},
            {"label": "1er rappel DTC (18 mois)", "db_name": "Rappel DTC1", "has_0_11": False, "has_12_59": True},
            {"label": "2ème rappel DTC (5 ans)", "db_name": "Rappel DTC2", "has_0_11": False, "has_5y": True},
        ]
        
        expected_vaxes = [t for t in table_structure if not t.get("is_header")]
        # Inject custom/archived vaccines dynamically
        known_db_names = {t["db_name"] for t in expected_vaxes}
        custom_vaxes = set()
        for rec in stats_data:
            mapped_name = ReportBuilder._map_vax_name(rec)
            if mapped_name not in known_db_names:
                custom_vaxes.add(mapped_name)
                
        if custom_vaxes:
            table_structure.append({"label": "Autres Vaccins (Archives)", "is_header": True})
            for cv in sorted(list(custom_vaxes)):
                new_entry = {"label": cv, "db_name": cv, "has_0_11": True, "has_12_59": True, "has_5y": True}
                table_structure.append(new_entry)
                expected_vaxes.append(new_entry)
                
        processed = ReportBuilder._process_data(stats_data, expected_vaxes)
        
        row_idx = 6
        for row in table_structure:
            is_stripe = (row_idx % 2 == 0)
            fill_style = striped_fill if is_stripe else PatternFill(fill_type=None)
            
            if row.get("is_header"):
                ws.merge_cells(f"A{row_idx}:K{row_idx}")
                cell = ws[f"A{row_idx}"]
                cell.value = row["label"]
                cell.font = sub_header_font
                cell.fill = sub_header_fill
                cell.border = thin_border
            else:
                db_name = row["db_name"]
                data = processed.get(db_name, {
                    "0-11m": {"M": 0, "F": 0, "Total": 0},
                    "12-59m": {"M": 0, "F": 0, "Total": 0},
                    "60m+": {"M": 0, "F": 0, "Total": 0}
                })
                
                has_12_59 = row.get("has_12_59", False)
                has_5y = row.get("has_5y", False)
                
                c_0_m = data["0-11m"]["M"] if row.get("has_0_11") else ""
                c_0_f = data["0-11m"]["F"] if row.get("has_0_11") else ""
                c_0_t = data["0-11m"]["Total"] if row.get("has_0_11") else ""
                
                c_12_m = data["12-59m"]["M"] if has_12_59 else ""
                c_12_f = data["12-59m"]["F"] if has_12_59 else ""
                c_12_t = data["12-59m"]["Total"] if has_12_59 else ""
                
                c_5_m = data["60m+"]["M"] if has_5y else ""
                c_5_f = data["60m+"]["F"] if has_5y else ""
                c_5_t = data["60m+"]["Total"] if has_5y else ""
                
                grand_total = data["0-11m"]["Total"] + data["12-59m"]["Total"] + data["60m+"]["Total"]
                grand_total_display = grand_total if grand_total > 0 else ""
                
                cols = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
                vals = [row['label'], c_0_m, c_0_f, c_0_t, c_12_m, c_12_f, c_12_t, c_5_m, c_5_f, c_5_t, grand_total_display]
                
                for c, v in zip(cols, vals):
                    ws[f"{c}{row_idx}"].value = v if v != 0 else ""
                    ws[f"{c}{row_idx}"].alignment = center_align if c != "A" else left_align
                    ws[f"{c}{row_idx}"].fill = fill_style
                    ws[f"{c}{row_idx}"].border = thin_border
                    
                    if c in ["D", "G", "J"]:
                        ws[f"{c}{row_idx}"].font = bold_font
                    elif c == "K":
                        ws[f"{c}{row_idx}"].font = total_font
                    else:
                        ws[f"{c}{row_idx}"].font = normal_font
                        
            row_idx += 1
            
        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]:
            ws.column_dimensions[col].width = 10
            
        return wb

    @staticmethod
    def generate_daily_breakdown_excel(engine, year, month, center_name, center_type):
        import calendar
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        y_int = int(year)
        m_int = int(month)
        _, num_days = calendar.monthrange(y_int, m_int)
        
        vax_rows = [
            ("HB0_24h", "HB0 (Dans les 24h)"), ("HB0_BCG", "HB0 (Plus tard)"), ("BCG", "BCG"),
            ("VPO0", "VPO 0"), ("VPO1", "VPO 1"), ("VPO2", "VPO 2"), ("VPO3", "VPO 3"),
            ("VPI", "VPI 1"), ("VPI2", "VPI 2"),
            ("Pentavalent 1", "Penta 1"), ("Pentavalent 2", "Penta 2"), ("Pentavalent 3", "Penta 3"),
            ("Rota1", "Rota 1"), ("Rota2", "Rota 2"), ("Rota3", "Rota 3"),
            ("PCV1", "Pneumo 1"), ("PCV2", "Pneumo 2"), ("PCV3", "Pneumo 3"), ("PCV4", "Pneumo 4"),
            ("RR1", "RR 1 (9 mois)"), ("RR2", "RR 2 (18 mois)"),
            ("VPO (18 Mois)", "Rappel VPO (18m)"), ("VPO (5 Ans)", "Rappel VPO (5 ans)"),
            ("Rappel DTC1", "Rappel DTC 1"), ("Rappel DTC2", "Rappel DTC 2")
        ]
        
        year_month_str = f"{year}-{month}%"
        stats_data = engine.db.get_detailed_export_stats(year_month_str)
        
        grid = {db_name: {d: 0 for d in range(1, num_days + 1)} for db_name, _ in vax_rows}
        day_totals = {d: 0 for d in range(1, num_days + 1)}
        
        for record in stats_data:
            mapped_name = ReportBuilder._map_vax_name(record)
            day = int(record["date_given"][-2:])
            
            if mapped_name not in grid:
                vax_rows.append((mapped_name, mapped_name))
                grid[mapped_name] = {d: 0 for d in range(1, num_days + 1)}
                
            if 1 <= day <= num_days:
                grid[mapped_name][day] += 1
                day_totals[day] += 1
                
        active_days = [d for d in range(1, num_days + 1) if day_totals[d] > 0]
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Journalier_{month}_{year}"
        
        # Styles
        title_font = Font(name="Segoe UI", size=14, bold=True, color="2C3E50")
        sub_title_font = Font(name="Segoe UI", size=10, italic=True)
        header_font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
        normal_font = Font(name="Segoe UI", size=9)
        bold_font = Font(name="Segoe UI", size=9, bold=True)
        total_font = Font(name="Segoe UI", size=10, bold=True, color="E74C3C")
        
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        striped_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        footer_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
        
        thin_border = Border(left=Side(style='thin', color="CCCCCC"), 
                             right=Side(style='thin', color="CCCCCC"), 
                             top=Side(style='thin', color="CCCCCC"), 
                             bottom=Side(style='thin', color="CCCCCC"))
        
        # Build Top
        end_col_num = len(active_days) + 2
        from openpyxl.utils import get_column_letter
        end_col = get_column_letter(end_col_num)
        
        ws.merge_cells(f"A1:{end_col}1")
        ws["A1"] = "REGISTRE JOURNALIER DE VACCINATION"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align
        
        ws.merge_cells(f"A2:{end_col}2")
        ws["A2"] = f"Mois: {month}/{year} | Etablissement: {center_name} ({center_type})"
        ws["A2"].font = sub_title_font
        ws["A2"].alignment = center_align
        
        # Headers
        ws["A4"] = "Vaccin"
        ws["A4"].font = header_font
        ws["A4"].alignment = center_align
        ws["A4"].fill = header_fill
        ws["A4"].border = thin_border
        
        for c_idx, d in enumerate(active_days):
            col_letter = get_column_letter(c_idx + 2)
            cell = ws[f"{col_letter}4"]
            cell.value = f"{d:02d}/{m_int:02d}"
            cell.font = header_font
            cell.alignment = center_align
            cell.fill = header_fill
            cell.border = thin_border
            
        cell_total_h = ws[f"{end_col}4"]
        cell_total_h.value = "Total"
        cell_total_h.font = header_font
        cell_total_h.alignment = center_align
        cell_total_h.fill = header_fill
        cell_total_h.border = thin_border
        
        # Rows
        row_idx = 5
        for db_name, label in vax_rows:
            is_stripe = (row_idx % 2 == 0)
            fill_style = striped_fill if is_stripe else PatternFill(fill_type=None)
            
            ws[f"A{row_idx}"].value = label
            ws[f"A{row_idx}"].font = bold_font
            ws[f"A{row_idx}"].alignment = left_align
            ws[f"A{row_idx}"].fill = fill_style
            ws[f"A{row_idx}"].border = thin_border
            
            row_tot = sum(grid[db_name][d] for d in range(1, num_days + 1))
            
            for c_idx, d in enumerate(active_days):
                val = grid[db_name][d]
                col_letter = get_column_letter(c_idx + 2)
                cell = ws[f"{col_letter}{row_idx}"]
                cell.value = val if val > 0 else "-"
                cell.font = normal_font if val == 0 else bold_font
                cell.alignment = center_align
                cell.fill = fill_style
                cell.border = thin_border
                
            cell_total = ws[f"{end_col}{row_idx}"]
            cell_total.value = row_tot if row_tot > 0 else "-"
            cell_total.font = total_font
            cell_total.alignment = center_align
            cell_total.fill = fill_style
            cell_total.border = thin_border
            
            row_idx += 1
            
        # Footer
        ws[f"A{row_idx}"].value = "Total Doses"
        ws[f"A{row_idx}"].font = bold_font
        ws[f"A{row_idx}"].alignment = left_align
        ws[f"A{row_idx}"].fill = footer_fill
        ws[f"A{row_idx}"].border = thin_border
        
        grand_total = sum(day_totals[d] for d in range(1, num_days + 1))
        
        for c_idx, d in enumerate(active_days):
            val = day_totals[d]
            col_letter = get_column_letter(c_idx + 2)
            cell = ws[f"{col_letter}{row_idx}"]
            cell.value = val if val > 0 else "-"
            cell.font = bold_font
            cell.alignment = center_align
            cell.fill = footer_fill
            cell.border = thin_border
            
        cell_grand = ws[f"{end_col}{row_idx}"]
        cell_grand.value = grand_total if grand_total > 0 else "-"
        cell_grand.font = total_font
        cell_grand.alignment = center_align
        cell_grand.fill = footer_fill
        cell_grand.border = thin_border
        
        # Dimensions
        ws.column_dimensions["A"].width = 20
        for c_idx in range(len(active_days) + 1):
            col_letter = get_column_letter(c_idx + 2)
            ws.column_dimensions[col_letter].width = 8
            
        return wb

    @staticmethod
    def _interpret_zscore(z, metric):
        if z is None:
            return ""
        if metric == "Poids":
            if z < -3: return "(Insuffisance sévère)"
            if z < -2: return "(Insuffisance pondérale)"
            if z > 2: return "(Poids élevé)"
            return "(Normal)"
        elif metric == "Taille":
            if z < -3: return "(Retard sévère)"
            if z < -2: return "(Retard de croissance)"
            if z > 3: return "(Très grande taille)"
            if z > 2: return "(Grande taille)"
            return "(Normal)"
        elif metric == "IMC":
            if z < -3: return "(Malnutrition sévère)"
            if z < -2: return "(Malnutrition)"
            if z > 3: return "(Obésité)"
            if z > 2: return "(Surpoids)"
            if z > 1: return "(Risque surpoids)"
            return "(Normal)"
        return ""

    @staticmethod
    def _process_nutrition_data(data):
        results = []
        for i, row in enumerate(data, 1):
            vaxes = row.get("vaccines", [])
            # Motif is either exact list of vaccines or "Mesures de croissance"
            motif = ", ".join(vaxes) if vaxes else "Mesures de croissance"
            
            # Map Vit A and Vit D
            vit_a1 = "Oui" if "Vit A1" in vaxes else ""
            vit_a2 = "Oui" if "Vit A2" in vaxes else ""
            vit_a3 = "Oui" if "Vit A3" in vaxes else ""
            vit_d1 = "Oui" if "Vit D1" in vaxes else ""
            vit_d2 = "Oui" if "Vit D2" in vaxes else ""
            
            # calculate age
            try:
                dob = datetime.strptime(row['dob'], "%Y-%m-%d").date()
                today = datetime.now().date()
                age_months = (today - dob).days / 30.44
                age_str = f"{age_months:.1f}"
            except Exception:
                age_str = "0.0"
                
            # Formatting Poids, Taille, IMC with their parsed Z-scores if available
            poids_str = f"{row['weight']}" if row.get('weight') is not None else ""
            taille_str = f"{row['height']}" if row.get('height') is not None else ""
            imc_str = f"{row['imc']}" if row.get('imc') is not None else ""
            
            if poids_str and row.get('z_w') is not None:
                interp = ReportBuilder._interpret_zscore(row['z_w'], "Poids")
                poids_str += f"\n{interp}"
                
            if taille_str and row.get('z_h') is not None:
                interp = ReportBuilder._interpret_zscore(row['z_h'], "Taille")
                taille_str += f"\n{interp}"
                
            if imc_str and row.get('z_i') is not None:
                interp = ReportBuilder._interpret_zscore(row['z_i'], "IMC")
                imc_str += f"\n{interp}"
                
            results.append({
                "record_date": row.get("record_date"),
                "n_ordre": i,
                "n_smi": row["patient_id"],
                "nom_prenom": row["name"],
                "dob": row["dob"],
                "age_mois": age_str,
                "sexe": "M" if row["sexe"] == "Garçon" or row["sexe"] == "M" else "F",
                "adresse": row["address"],
                "motif": motif,
                "poids": poids_str,
                "taille": taille_str,
                "imc": imc_str,
                "vit_d1": vit_d1,
                "vit_d2": vit_d2,
                "vit_a1": vit_a1,
                "vit_a2": vit_a2,
                "vit_a3": vit_a3,
            })
        return results

    @staticmethod
    def generate_nutrition_html(data, date_str, center_name, center_type):
        d_fmt = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d/%m/%Y")
        processed_data = ReportBuilder._process_nutrition_data(data)
        
        html = f"""
        <html>
        <head>
        <style>
            @page {{ margin: 10mm; size: landscape; }}
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; font-size: 8pt; color: #333; margin: 0; padding: 0; }}
            h1 {{ text-align: center; font-size: 16pt; margin-bottom: 5px; color: #2c3e50; text-transform: uppercase; letter-spacing: 1px; }}
            .header-info {{ 
                text-align: center; font-size: 10pt; margin-bottom: 15px; color: #555;
                background-color: #ecf0f1; border-radius: 6px; padding: 8px; max-width: 600px; margin-left: auto; margin-right: auto;
                border-left: 5px solid #27ae60;
            }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; box-shadow: 0 1px 4px rgba(0,0,0,0.1); border-radius: 4px; overflow: hidden; page-break-inside: auto; }}
            th, td {{ border: 1px solid #ccc; padding: 4px; text-align: center; vertical-align: middle; page-break-inside: avoid; }}
            th {{ background-color: #34495e; color: #fff; font-weight: bold; font-size: 8pt; }}
            .small-col {{ width: 3%; }}
            thead {{ display: table-header-group; }}
            tr {{ page-break-inside: avoid; page-break-after: auto; }}
            tr:nth-child(even) {{ background-color: #fafafa; }}
            tr:hover {{ background-color: #e8f4f8; }}
        </style>
        </head>
        <body>
            <h1>REGISTRE INTÉGRÉ DE NUTRITION</h1>
            <div class="header-info">
                Journée du : <strong>{d_fmt}</strong><br>
                Etablissement : <strong>{center_name}</strong> ({center_type})
            </div>
        """
            
        # Chunk the data into pages to manually force headers and page breaks
        rows_per_page = 12
        pages = [processed_data[i:i + rows_per_page] for i in range(0, max(1, len(processed_data)), rows_per_page)]
        
        for p_idx, page_rows in enumerate(pages):
            if p_idx > 0:
                html += '<div style="page-break-before: always;"></div>'
                
            html += f"""
            <table>
                <thead>
                    <tr>
                        <th rowspan="3" class="small-col">N&#176; Ordre</th>
                        <th rowspan="3" class="small-col">N&#176; SMI</th>
                        <th rowspan="3" style="width: 12%;">Nom, Prénom</th>
                        <th rowspan="3" style="width: 10%;">Date de naissance<br>ou âge en mois</th>
                        <th rowspan="3" class="small-col">Sexe<br>(M/F)</th>
                        <th rowspan="3" style="width: 15%;">Adresse et N° tél</th>
                        <th rowspan="3" style="width: 25%;">Motif de la consultation</th>
                        <th colspan="3" rowspan="2">Mesures de croissance</th>
                        <th colspan="5">Supplémentation</th>
                    </tr>
                    <tr>
                        <th colspan="2">Vit D</th>
                        <th colspan="3">Vit A</th>
                    </tr>
                    <tr>
                        <th>Poids<br>(kg)</th>
                        <th>Taille<br>(cm)</th>
                        <th>IMC</th>
                        <th>1ère prise</th>
                        <th>2ème prise</th>
                        <th>100.000 UI</th>
                        <th colspan="2">200.000 UI</th>
                    </tr>
                </thead>
                <tbody>
            """
            
            for row in page_rows:
                html += f"""
                    <tr>
                        <td>{row['n_ordre']}</td>
                        <td>{row['n_smi']}</td>
                        <td>{row['nom_prenom']}</td>
                        <td>{row['dob']}<br>{row['age_mois']} m</td>
                        <td>{row['sexe']}</td>
                        <td>{row['adresse']}</td>
                        <td>{row['motif']}</td>
                        <td>{row['poids']}</td>
                        <td>{row['taille']}</td>
                        <td>{row['imc']}</td>
                        <td>{row['vit_d1']}</td>
                        <td>{row['vit_d2']}</td>
                        <td>{row['vit_a1']}</td>
                        <td>{row['vit_a2']}</td>
                        <td>{row['vit_a3']}</td>
                    </tr>
                """
                
            html += """
                </tbody>
            </table>
            """
            
        html += """
        </body>
        </html>
        """
        return html

    @staticmethod
    def generate_nutrition_excel(data, date_str, center_name, center_type):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        d_fmt = datetime.strptime(date_str, "%Y-%m-%d").strftime("%d/%m/%Y")
        processed_data = ReportBuilder._process_nutrition_data(data)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Nutrition_{date_str}"
        
        # Styles
        title_font = Font(name="Segoe UI", size=14, bold=True, color="2C3E50")
        sub_title_font = Font(name="Segoe UI", size=10, italic=True)
        header_font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
        normal_font = Font(name="Segoe UI", size=9)
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        striped_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        
        thin_border = Border(left=Side(style='thin', color="CCCCCC"), 
                             right=Side(style='thin', color="CCCCCC"), 
                             top=Side(style='thin', color="CCCCCC"), 
                             bottom=Side(style='thin', color="CCCCCC"))
                             
        # Titles
        ws.merge_cells("A1:O1")
        ws["A1"] = "REGISTRE INTÉGRÉ DE NUTRITION"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align
        
        ws.merge_cells("A2:O2")
        ws["A2"] = f"Journée du: {d_fmt} | Etablissement: {center_name} ({center_type})"
        ws["A2"].font = sub_title_font
        ws["A2"].alignment = center_align
        
        # Headers
        ws.merge_cells("A4:A6")
        ws["A4"] = "N° Ordre"
        ws.merge_cells("B4:B6")
        ws["B4"] = "N° SMI"
        ws.merge_cells("C4:C6")
        ws["C4"] = "Nom, Prénom"
        ws.merge_cells("D4:D6")
        ws["D4"] = "Date de naissance\nou âge en mois"
        ws.merge_cells("E4:E6")
        ws["E4"] = "Sexe\n(M/F)"
        ws.merge_cells("F4:F6")
        ws["F4"] = "Adresse et N° tél"
        ws.merge_cells("G4:G6")
        ws["G4"] = "Motif de la consultation"
        
        ws.merge_cells("H4:J5")
        ws["H4"] = "Mesures de croissance"
        
        ws.merge_cells("K4:O4")
        ws["K4"] = "Supplémentation"
        
        ws.merge_cells("K5:L5")
        ws["K5"] = "Vit D"
        ws.merge_cells("M5:O5")
        ws["M5"] = "Vit A"
        
        ws["H6"] = "Poids (kg)"
        ws["I6"] = "Taille (cm)"
        ws["J6"] = "IMC"
        ws["K6"] = "1ère prise"
        ws["L6"] = "2ème prise"
        ws["M6"] = "100.000 UI\n(1ère prise)"
        ws.merge_cells("N6:O6")
        ws["N6"] = "200.000 UI\n(2ème/3ème prise)"
        
        # Apply header styles
        for r in range(4, 7):
            for c in range(1, 16):
                col_letter = get_column_letter(c)
                cell = ws[f"{col_letter}{r}"]
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = header_fill
                cell.border = thin_border
                
        # Data Rows
        row_idx = 7
        for row in processed_data:
            is_stripe = (row_idx % 2 == 0)
            fill_style = striped_fill if is_stripe else PatternFill(fill_type=None)
            
            vals = [
                row['n_ordre'], row['n_smi'], row['nom_prenom'], 
                f"{row['dob']}\n({row['age_mois']}m)", row['sexe'], 
                row['adresse'], row['motif'], 
                row['poids'], row['taille'], row['imc'],
                row['vit_d1'], row['vit_d2'], row['vit_a1'], row['vit_a2'], row['vit_a3']
            ]
            
            for c_idx, val in enumerate(vals):
                col_letter = get_column_letter(c_idx + 1)
                cell = ws[f"{col_letter}{row_idx}"]
                cell.value = val
                cell.font = normal_font
                cell.fill = fill_style
                cell.border = thin_border
                
                if c_idx in [2, 5, 6]:
                    cell.alignment = left_align
                else:
                    cell.alignment = center_align
                    
            row_idx += 1
            
        # Adjust column widths
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 6
        ws.column_dimensions["F"].width = 25
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].width = 10
        ws.column_dimensions["J"].width = 10
        ws.column_dimensions["K"].width = 10
        ws.column_dimensions["L"].width = 10
        ws.column_dimensions["M"].width = 12
        ws.column_dimensions["N"].width = 10
        ws.column_dimensions["O"].width = 10
        
        return wb

    @staticmethod
    def generate_multi_nutrition_html(data, dates, center_name, center_type):
        html = f"""
        <html>
        <head>
        <style>
            @page {{ margin: 10mm; size: landscape; }}
            body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; font-size: 8pt; color: #333; margin: 0; padding: 0; }}
            h1 {{ text-align: center; font-size: 16pt; margin-bottom: 5px; color: #2c3e50; text-transform: uppercase; letter-spacing: 1px; }}
            .header-info {{ 
                text-align: center; font-size: 10pt; margin-bottom: 15px; color: #555;
                background-color: #ecf0f1; border-radius: 6px; padding: 8px; max-width: 600px; margin-left: auto; margin-right: auto;
                border-left: 5px solid #27ae60;
            }}
            table {{ width: 100%; border-collapse: collapse; margin-top: 10px; box-shadow: 0 1px 4px rgba(0,0,0,0.1); border-radius: 4px; overflow: hidden; page-break-inside: auto; }}
            th, td {{ border: 1px solid #ccc; padding: 4px; text-align: center; vertical-align: middle; page-break-inside: avoid; }}
            th {{ background-color: #34495e; color: #fff; font-weight: bold; font-size: 8pt; }}
            .small-col {{ width: 3%; }}
            .date-header {{ background-color: #dcdde1; color: #2c3e50; font-weight: bold; font-size: 10pt; text-align: center; }}
            thead {{ display: table-header-group; }}
            tr {{ page-break-inside: avoid; page-break-after: auto; }}
            tr:nth-child(even) {{ background-color: #fafafa; }}
            tr:hover {{ background-color: #e8f4f8; }}
        </style>
        </head>
        <body>
            <h1>REGISTRE INTÉGRÉ DE NUTRITION</h1>
            <div class="header-info">
                Registre Aggrégé (Multi-Dates)<br>
                Etablissement : <strong>{center_name}</strong> ({center_type})
            </div>
        """
        # Extract flat chronological list of all rows across all dates
        grouped = {}
        processed_data = ReportBuilder._process_nutrition_data(data)
        for row in processed_data:
            dt = row.get("record_date")
            if dt not in grouped: grouped[dt] = []
            grouped[dt].append(row)
            
        flat_rows = []
        for d in dates:
            if d not in grouped: continue
            flat_rows.append({"is_date_header": True, "date_str": d})
            for row in grouped[d]:
                row["is_date_header"] = False
                flat_rows.append(row)
                
        # Chunk the data into pages to manually force headers and page breaks
        rows_per_page = 13 # Adjusted for spacing
        pages = [flat_rows[i:i + rows_per_page] for i in range(0, max(1, len(flat_rows)), rows_per_page)]
        
        for p_idx, page_rows in enumerate(pages):
            if p_idx > 0:
                html += '<div style="page-break-before: always;"></div>'
                
            html += f"""
            <table>
                <thead>
                    <tr>
                        <th rowspan="3" class="small-col">N&#176; Ordre</th>
                        <th rowspan="3" class="small-col">N&#176; SMI</th>
                        <th rowspan="3" style="width: 12%;">Nom, Prénom</th>
                        <th rowspan="3" style="width: 10%;">Date de naissance<br>ou âge en mois</th>
                        <th rowspan="3" class="small-col">Sexe<br>(M/F)</th>
                        <th rowspan="3" style="width: 15%;">Adresse et N&#176; tél</th>
                        <th rowspan="3" style="width: 25%;">Motif de la consultation</th>
                        <th colspan="3" rowspan="2">Mesures de croissance</th>
                        <th colspan="5">Supplémentation</th>
                    </tr>
                    <tr>
                        <th colspan="2">Vit D</th>
                        <th colspan="3">Vit A</th>
                    </tr>
                    <tr>
                        <th>Poids<br>(kg)</th>
                        <th>Taille<br>(cm)</th>
                        <th>IMC</th>
                        <th>1ère prise</th>
                        <th>2ème prise</th>
                        <th>100.000 UI</th>
                        <th colspan="2">200.000 UI</th>
                    </tr>
                </thead>
                <tbody>
            """
            
            for row in page_rows:
                if row.get("is_date_header"):
                    d_fmt = datetime.strptime(row["date_str"], "%Y-%m-%d").strftime("%d/%m/%Y")
                    html += f"""
                        <tr>
                            <td colspan="15" class="date-header">Journée du : {d_fmt}</td>
                        </tr>
                    """
                else:
                    html += f"""
                        <tr>
                            <td>{row['n_ordre']}</td>
                            <td>{row['n_smi']}</td>
                            <td>{row['nom_prenom']}</td>
                            <td>{row['dob']}<br>{row['age_mois']} m</td>
                            <td>{row['sexe']}</td>
                            <td>{row['adresse']}</td>
                            <td>{row['motif']}</td>
                            <td>{row['poids']}</td>
                            <td>{row['taille']}</td>
                            <td>{row['imc']}</td>
                            <td>{row['vit_d1']}</td>
                            <td>{row['vit_d2']}</td>
                            <td>{row['vit_a1']}</td>
                            <td>{row['vit_a2']}</td>
                            <td>{row['vit_a3']}</td>
                        </tr>
                    """
                    
            html += """
                </tbody>
            </table>
            """
            
        html += """
        </body>
        </html>
        """
        return html

    @staticmethod
    def generate_multi_nutrition_excel(data, dates, center_name, center_type):
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Nutrition_Agrege"
        
        title_font = Font(name="Segoe UI", size=14, bold=True, color="2C3E50")
        sub_title_font = Font(name="Segoe UI", size=10, italic=True)
        header_font = Font(name="Segoe UI", size=9, bold=True, color="FFFFFF")
        normal_font = Font(name="Segoe UI", size=9)
        date_header_font = Font(name="Segoe UI", size=11, bold=True, color="2C3E50")
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        striped_fill = PatternFill(start_color="F9F9F9", end_color="F9F9F9", fill_type="solid")
        date_header_fill = PatternFill(start_color="DCDDE1", end_color="DCDDE1", fill_type="solid")
        
        thin_border = Border(left=Side(style='thin', color="CCCCCC"), 
                             right=Side(style='thin', color="CCCCCC"), 
                             top=Side(style='thin', color="CCCCCC"), 
                             bottom=Side(style='thin', color="CCCCCC"))
                             
        ws.merge_cells("A1:O1")
        ws["A1"] = "REGISTRE INTÉGRÉ DE NUTRITION"
        ws["A1"].font = title_font
        ws["A1"].alignment = center_align
        
        ws.merge_cells("A2:O2")
        ws["A2"] = f"Registre Aggrégé (Multi-Dates) | Etablissement: {center_name} ({center_type})"
        ws["A2"].font = sub_title_font
        ws["A2"].alignment = center_align
        
        ws.merge_cells("A4:A6")
        ws["A4"] = "N° Ordre"
        ws.merge_cells("B4:B6")
        ws["B4"] = "N° SMI"
        ws.merge_cells("C4:C6")
        ws["C4"] = "Nom, Prénom"
        ws.merge_cells("D4:D6")
        ws["D4"] = "Date de naissance\nou âge en mois"
        ws.merge_cells("E4:E6")
        ws["E4"] = "Sexe\n(M/F)"
        ws.merge_cells("F4:F6")
        ws["F4"] = "Adresse et N° tél"
        ws.merge_cells("G4:G6")
        ws["G4"] = "Motif de la consultation"
        
        ws.merge_cells("H4:J5")
        ws["H4"] = "Mesures de croissance"
        
        ws.merge_cells("K4:O4")
        ws["K4"] = "Supplémentation"
        
        ws.merge_cells("K5:L5")
        ws["K5"] = "Vit D"
        ws.merge_cells("M5:O5")
        ws["M5"] = "Vit A"
        
        ws["H6"] = "Poids (kg)"
        ws["I6"] = "Taille (cm)"
        ws["J6"] = "IMC"
        ws["K6"] = "1ère prise"
        ws["L6"] = "2ème prise"
        ws["M6"] = "100.000 UI\n(1ère prise)"
        ws.merge_cells("N6:O6")
        ws["N6"] = "200.000 UI\n(2ème/3ème prise)"
        
        for r in range(4, 7):
            for c in range(1, 16):
                col_letter = get_column_letter(c)
                cell = ws[f"{col_letter}{r}"]
                cell.font = header_font
                cell.alignment = center_align
                cell.fill = header_fill
                cell.border = thin_border
                
        # Group and Draw Rows
        grouped = {}
        processed_data = ReportBuilder._process_nutrition_data(data)
        for row in processed_data:
            dt = row.get("record_date")
            if dt not in grouped: grouped[dt] = []
            grouped[dt].append(row)
            
        row_idx = 7
        for d in dates:
            if d not in grouped: continue
            
            d_fmt = datetime.strptime(d, "%Y-%m-%d").strftime("%d/%m/%Y")
            
            # Print inline date header
            ws.merge_cells(f"A{row_idx}:O{row_idx}")
            ws[f"A{row_idx}"] = f"Journée du : {d_fmt}"
            ws[f"A{row_idx}"].font = date_header_font
            ws[f"A{row_idx}"].alignment = center_align
            ws[f"A{row_idx}"].fill = date_header_fill
            ws[f"A{row_idx}"].border = thin_border
            for c in range(2, 16):
                ws[f"{get_column_letter(c)}{row_idx}"].border = thin_border
            row_idx += 1
            
            # Print rows for this date
            for row in grouped[d]:
                is_stripe = (row_idx % 2 == 0)
                fill_style = striped_fill if is_stripe else PatternFill(fill_type=None)
                
                vals = [
                    row['n_ordre'], row['n_smi'], row['nom_prenom'], 
                    f"{row['dob']}\n({row['age_mois']}m)", row['sexe'], 
                    row['adresse'], row['motif'], 
                    row['poids'], row['taille'], row['imc'],
                    row['vit_d1'], row['vit_d2'], row['vit_a1'], row['vit_a2'], row['vit_a3']
                ]
                
                for c_idx, val in enumerate(vals):
                    col_letter = get_column_letter(c_idx + 1)
                    cell = ws[f"{col_letter}{row_idx}"]
                    cell.value = val
                    cell.font = normal_font
                    cell.fill = fill_style
                    cell.border = thin_border
                    
                    if c_idx in [2, 5, 6]:
                        cell.alignment = left_align
                    else:
                        cell.alignment = center_align
                        
                row_idx += 1

        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 6
        ws.column_dimensions["F"].width = 25
        ws.column_dimensions["G"].width = 20
        ws.column_dimensions["H"].width = 10
        ws.column_dimensions["I"].width = 10
        ws.column_dimensions["J"].width = 10
        ws.column_dimensions["K"].width = 10
        ws.column_dimensions["L"].width = 10
        ws.column_dimensions["M"].width = 12
        ws.column_dimensions["N"].width = 10
        ws.column_dimensions["O"].width = 10
        
        return wb
